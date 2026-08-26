import copy
import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.generate_xlsx_output import (
    build_workbook,
    compute_metrics,
    load_answer_key,
    load_run_outputs,
    load_source_accruals,
)


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"


class EvaluationIntegrityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.outputs, cls.problems = load_run_outputs(OUTPUTS)
        _counts, cls.answer_entries = load_answer_key(OUTPUTS / "answer_key.json.gz")

    def test_v1_defect_is_reported_honestly(self):
        metrics = compute_metrics(self.outputs, self.answer_entries, self.problems)
        self.assertEqual(metrics["overall_status"], "Needs Repair")
        self.assertAlmostEqual(metrics["unresolvable_precision"], 25 / 34)
        self.assertEqual(metrics["unresolvable_recall"], 1.0)
        self.assertEqual(metrics["terminal_misroute_count"], 9)
        self.assertAlmostEqual(metrics["terminal_misroute_value_gbp"], 3212.35)
        self.assertEqual(metrics["evidence_allocation_coverage"], 0.0)
        self.assertAlmostEqual(metrics["resolved_value_gbp"], 84525.13)
        self.assertAlmostEqual(metrics["v1_plausibility_only_value_gbp"], 47461.98)

    def test_missing_output_can_never_be_reported_as_zero(self):
        metrics = compute_metrics(
            self.outputs,
            self.answer_entries,
            ["missing:data_quality_issues.json"],
        )
        self.assertEqual(metrics["overall_status"], "Incomplete Run")

    def test_zero_accepted_classifications_is_not_excellent(self):
        outputs = copy.deepcopy(self.outputs)
        outputs["needs_review"].extend(outputs["classified_verified"])
        outputs["classified_verified"] = []
        metrics = compute_metrics(outputs, self.answer_entries, [])
        self.assertIsNone(metrics["classification_precision"])
        self.assertNotEqual(metrics["overall_status"], "Ready for Demo")

    def test_workbook_contains_finance_and_evaluation_views(self):
        metrics = compute_metrics(self.outputs, self.answer_entries, self.problems)
        outputs = copy.deepcopy(self.outputs)
        outputs["needs_review"] = []
        outputs["data_quality_issue"] = []
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "report.xlsx"
            build_workbook(outputs, metrics, destination)
            self.assertTrue(destination.exists())
            self.assertGreater(destination.stat().st_size, 10_000)
            workbook = load_workbook(destination, read_only=False, data_only=False)
            self.assertEqual(workbook["Needs_Review"]["A1"].value, "Priority")
            self.assertEqual(
                workbook["Needs_Review"]["A2"].value,
                "No deductions need human classification review in this run.",
            )
            self.assertEqual(workbook["Data_Quality"]["A1"].value, "Priority")
            self.assertEqual(
                workbook["Data_Quality"]["A2"].value,
                "No source-data quality issues were detected in this run.",
            )
            self.assertEqual(
                workbook["Start_Here"]["A38"].value,
                "CLASSIFICATION ROUTE DISCLOSURE",
            )

    def test_v2_workbook_discloses_route_sources_and_human_authority(self):
        output_dir = ROOT / "outputs" / "v2"
        outputs, problems = load_run_outputs(output_dir)
        _counts, entries = load_answer_key(ROOT / "fixtures" / "v2" / "answer_key.json.gz")
        accruals, accrual_problems = load_source_accruals(
            ROOT / "fixtures" / "v2" / "invoice_accruals.csv"
        )
        metrics = compute_metrics(outputs, entries, problems, accruals, accrual_problems)
        self.assertEqual(metrics["accepted_classification_method_counts"], {"configured_alias": 75})
        self.assertEqual(metrics["unresolvable_method_counts"], {"ai_proposal": 25})
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "v2-report.xlsx"
            build_workbook(outputs, metrics, destination)
            workbook = load_workbook(destination, read_only=False, data_only=False)
            summary = workbook["Start_Here"]
            self.assertEqual(summary["B41"].value, 75)
            self.assertEqual(summary["B42"].value, 0)
            self.assertEqual(summary["B43"].value, 25)
            self.assertEqual(summary["D8"].fill.fgColor.rgb, "002E7D32")
            classified = workbook["Classified_Evidence"]
            headers = {cell.value: cell.column for cell in classified[1]}
            self.assertEqual(
                classified.cell(2, headers["Status"]).value,
                "Accepted — evidence allocated",
            )
            self.assertIn(
                "human-controlled",
                classified.cell(2, headers["Next Action"]).value,
            )

            unsafe_metrics = copy.deepcopy(metrics)
            unsafe_metrics["terminal_misroute_count"] = 1
            unsafe_destination = Path(directory) / "unsafe-report.xlsx"
            build_workbook(outputs, unsafe_metrics, unsafe_destination)
            unsafe_workbook = load_workbook(unsafe_destination, read_only=False, data_only=False)
            self.assertEqual(
                unsafe_workbook["Start_Here"]["D8"].fill.fgColor.rgb,
                "00C00000",
            )

    def test_source_and_model_text_cannot_become_excel_formulas(self):
        outputs = {key: [] for key in self.outputs}
        outputs["needs_review"] = [{
            "deduction_id": "=HYPERLINK(\"https://example.invalid\",\"open\")",
            "transaction_date": "2026-08-25",
            "vendor_name": "@SUM(1,1)",
            "amount": 100,
            "llm_bucket": "Promotional Accrual",
            "llm_reasoning": "+cmd|' /C calc'!A0",
            "_evidence_note": "-1+1",
        }]
        metrics = compute_metrics(outputs, {}, ["synthetic:formula-neutralisation-test"])
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "formula-safe.xlsx"
            build_workbook(outputs, metrics, destination)
            workbook = load_workbook(destination, read_only=False, data_only=False)
            details = workbook["Needs_Review"]
            self.assertEqual(details["B2"].data_type, "s")
            self.assertTrue(details["B2"].value.startswith("'="))
            self.assertEqual(details["D2"].data_type, "s")
            self.assertTrue(details["D2"].value.startswith("'@"))
            self.assertEqual(details["K2"].data_type, "s")
            self.assertTrue(details["K2"].value.startswith("'-"))
            worklist = workbook["Priority_Worklist"]
            self.assertEqual(worklist["C2"].data_type, "s")
            self.assertTrue(worklist["C2"].value.startswith("'="))
            self.assertEqual(worklist["E2"].data_type, "s")
            self.assertTrue(worklist["E2"].value.startswith("'@"))
            self.assertEqual(worklist["H2"].data_type, "s")
            self.assertTrue(worklist["H2"].value.startswith("'-"))

    def test_duplicate_or_overallocated_source_balance_blocks_ready_status(self):
        outputs, problems = load_run_outputs(ROOT / "outputs" / "v2")
        _counts, entries = load_answer_key(ROOT / "fixtures" / "v2" / "answer_key.json.gz")
        accruals, accrual_problems = load_source_accruals(
            ROOT / "fixtures" / "v2" / "invoice_accruals.csv"
        )
        duplicated = copy.deepcopy(accruals)
        duplicated.append(copy.deepcopy(duplicated[0]))
        metrics = compute_metrics(outputs, entries, problems, duplicated, accrual_problems)
        self.assertEqual(metrics["overall_status"], "Incomplete Run")
        self.assertFalse(metrics["source_allocation_integrity"])

        overallocated_outputs = copy.deepcopy(outputs)
        overallocated_outputs["classified_verified"].append(
            copy.deepcopy(overallocated_outputs["classified_verified"][0])
        )
        overallocated_outputs["classified_verified"][-1]["deduction_id"] = "synthetic-overallocation"
        entries_with_control = copy.deepcopy(entries)
        entries_with_control["synthetic-overallocation"] = {
            "deduction_id": "synthetic-overallocation",
            "seed_category": "classifiable",
            "true_bucket": overallocated_outputs["classified_verified"][-1]["llm_bucket"],
        }
        row = overallocated_outputs["classified_verified"][-1]
        row["_allocated_amount"] = 999999.0
        row["_evidence_balance_before"] = 999999.0
        row["_evidence_balance_after"] = 0.0
        metrics = compute_metrics(
            overallocated_outputs, entries_with_control, problems, accruals, accrual_problems
        )
        self.assertEqual(metrics["overall_status"], "Needs Repair")
        self.assertFalse(metrics["source_allocation_integrity"])
        self.assertTrue(metrics["overallocated_accrual_ids"])

    def test_negative_allocation_and_reused_transaction_source_cannot_pass(self):
        outputs, problems = load_run_outputs(ROOT / "outputs" / "v2")
        _counts, entries = load_answer_key(ROOT / "fixtures" / "v2" / "answer_key.json.gz")
        accruals, accrual_problems = load_source_accruals(
            ROOT / "fixtures" / "v2" / "invoice_accruals.csv"
        )

        negative = copy.deepcopy(outputs)
        negative["classified_verified"][0]["_allocated_amount"] *= -1
        metrics = compute_metrics(negative, entries, problems, accruals, accrual_problems)
        self.assertEqual(metrics["overall_status"], "Needs Repair")
        self.assertLess(metrics["evidence_allocation_coverage"], 1.0)
        self.assertFalse(metrics["source_allocation_integrity"])

        reused = copy.deepcopy(outputs)
        reused["auto_matched"][1]["matched_accrual_id"] = reused["auto_matched"][0]["matched_accrual_id"]
        metrics = compute_metrics(reused, entries, problems, accruals, accrual_problems)
        self.assertGreaterEqual(metrics["auto_match_pair_correctness"], 0.99)
        self.assertEqual(metrics["overall_status"], "Needs Repair")
        self.assertFalse(metrics["transaction_match_integrity"])
        self.assertTrue(metrics["reused_transaction_match_ids"])

        altered_source = copy.deepcopy(accruals)
        matched_id = outputs["auto_matched"][0]["matched_accrual_id"]
        source = next(row for row in altered_source if row["accrual_id"] == matched_id)
        source["amount"] = str(float(source["amount"]) + 50)
        metrics = compute_metrics(outputs, entries, problems, altered_source, accrual_problems)
        self.assertEqual(metrics["overall_status"], "Needs Repair")
        self.assertFalse(metrics["transaction_match_integrity"])
        self.assertIn(matched_id, metrics["invalid_transaction_match_ids"])

    def test_source_ledger_loader_rejects_numeric_prefixes_and_nonfinite_amounts(self):
        for malformed in ("100GBP", "NaN", "Infinity", "1e3", "0", "-10"):
            with self.subTest(malformed=malformed), tempfile.TemporaryDirectory() as directory:
                path = Path(directory) / "accruals.csv"
                with path.open("w", encoding="utf-8", newline="") as handle:
                    writer = csv.DictWriter(
                        handle, fieldnames=["accrual_id", "amount", "evidence_scope"]
                    )
                    writer.writeheader()
                    writer.writerow({
                        "accrual_id": "A1", "amount": malformed,
                        "evidence_scope": "programme_pool",
                    })
                _rows, problems = load_source_accruals(path)
                self.assertTrue(problems)


if __name__ == "__main__":
    unittest.main()
