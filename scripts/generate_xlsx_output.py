"""Strictly evaluate a deduction run and build its finance-facing workbook.

V2 never treats a missing output as zero, reports precision and recall
separately, and leads with monetary outcomes plus a ranked human worklist.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from .deduction_rules import days_between, to_cents, vendor_score
except ImportError:  # Direct CLI execution from the scripts directory.
    from deduction_rules import days_between, to_cents, vendor_score

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs"
DEFAULT_ANSWER_KEY = DEFAULT_OUTPUT_DIR / "answer_key.json.gz"
DEFAULT_ACCRUAL_CSV = DEFAULT_OUTPUT_DIR / "invoice_accruals.csv"
DEFAULT_XLSX = DEFAULT_OUTPUT_DIR / "deduction_classification_output.xlsx"
DEFAULT_METRICS_JSON = DEFAULT_OUTPUT_DIR / "evaluation_metrics.json"

OUTCOME_FILES = {
    "auto_matched": "auto_matched.json",
    "classified_verified": "classified_verified.json",
    "needs_review": "needs_review.json",
    "unresolvable": "unresolvable.json",
    "data_quality_issue": "data_quality_issues.json",
}
OUTCOME_LABELS = {
    "auto_matched": "Auto-matched",
    "classified_verified": "Classified with allocated evidence",
    "needs_review": "Needs review",
    "unresolvable": "Unresolvable from supplied data",
    "data_quality_issue": "Data-quality issue",
}

NAVY, BLUE, GREEN, RED = "172B4D", "2E75B6", "2E7D32", "C00000"
LIGHT_GREEN, LIGHT_AMBER, LIGHT_RED = "E2F0D9", "FFF2CC", "FCE4D6"
LIGHT_BLUE, GREY, WHITE = "D9EAF7", "E7E6E6", "FFFFFF"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color=WHITE, bold=True)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
STATUS_FILLS = {
    "Ready for Demo": PatternFill("solid", fgColor=LIGHT_GREEN),
    "Needs Repair": PatternFill("solid", fgColor=LIGHT_RED),
    "Incomplete Run": PatternFill("solid", fgColor=LIGHT_AMBER),
    "Not Assessable": PatternFill("solid", fgColor=GREY),
}
THIN_GREY = Side(style="thin", color="D9E1F2")


def rate(numerator: int, denominator: int) -> float | None:
    return numerator / denominator if denominator else None


def pct(value: float | None) -> str:
    return "n/a" if value is None else f"{value:.1%}"


def record_value(record: dict[str, Any]) -> float:
    try:
        value = abs(float(record.get("amount") or 0))
        return value if math.isfinite(value) else 0.0
    except (TypeError, ValueError):
        return 0.0


def has_reconciling_allocation(row: dict[str, Any]) -> bool:
    try:
        allocated = float(row.get("_allocated_amount"))
        before = float(row.get("_evidence_balance_before"))
        after = float(row.get("_evidence_balance_after"))
    except (TypeError, ValueError):
        return False
    return (
        bool(row.get("_evidence_accrual_id"))
        and all(math.isfinite(value) for value in (allocated, before, after))
        and allocated > 0
        and after >= 0
        and abs(before - after - allocated) <= 0.011
    )


def load_json_list(path: Path) -> tuple[list[dict[str, Any]], str | None]:
    """Load a required branch without silently turning absence into a clean zero."""
    if not path.exists():
        return [], f"missing:{path.name}"
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return [], f"invalid:{path.name}:{exc}"
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list) or any(not isinstance(item, dict) for item in data):
        return [], f"invalid:{path.name}:expected a JSON array of objects"
    return data, None


def load_run_outputs(output_dir: Path) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    outputs, problems = {}, []
    for outcome, filename in OUTCOME_FILES.items():
        records, problem = load_json_list(output_dir / filename)
        outputs[outcome] = records
        if problem:
            problems.append(problem)
    return outputs, problems


def load_answer_key(path: Path) -> tuple[dict[str, int], dict[str, dict[str, Any]]]:
    with gzip.open(path, "rt", encoding="utf-8") as handle:
        data = json.load(handle)
    return data["counts"], {entry["deduction_id"]: entry for entry in data["entries"]}


def load_source_accruals(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Load the authoritative balances and make identity problems explicit."""
    if not path.exists():
        return [], [f"missing:{path.name}"]
    try:
        with path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
    except OSError as exc:
        return [], [f"invalid:{path.name}:{exc}"]
    required = {
        "accrual_id", "accrual_date", "vendor_name", "bucket", "amount",
        "evidence_scope",
    }
    if not rows or not required.issubset(rows[0]):
        return rows, [
            f"invalid:{path.name}:required columns are "
            "accrual_id, accrual_date, vendor_name, bucket, amount and evidence_scope"
        ]
    ids = [str(row.get("accrual_id") or "").strip() for row in rows]
    counts = Counter(ids)
    problems = []
    if any(not identifier for identifier in ids):
        problems.append(f"invalid:{path.name}:blank accrual_id")
    duplicates = sorted(identifier for identifier, count in counts.items() if identifier and count > 1)
    if duplicates:
        problems.append(f"invalid:{path.name}:duplicate accrual_id count={len(duplicates)}")
    for row in rows:
        amount_text = str(row.get("amount") or "").strip()
        if not re.fullmatch(r"\+?(?:\d+(?:\.\d{1,2})?|\.\d{1,2})", amount_text):
            problems.append(f"invalid:{path.name}:non-positive or invalid amount")
            break
        amount = float(amount_text)
        cents = round(amount * 100)
        if not math.isfinite(amount) or cents <= 0:
            problems.append(f"invalid:{path.name}:non-positive or invalid amount")
            break
        if row.get("evidence_scope") not in {"transaction_match", "programme_pool"}:
            problems.append(f"invalid:{path.name}:invalid evidence_scope")
            break
    return rows, problems


def dedupe_token_totals(records: list[dict[str, Any]]) -> dict[str, int]:
    calls = {}
    for record in records:
        batch_ref = record.get("_llm_batch_ref")
        if not batch_ref or batch_ref in calls:
            continue
        calls[batch_ref] = {
            key: int(record.get(key) or 0)
            for key in (
                "input_tokens", "output_tokens", "cache_creation_input_tokens",
                "cache_read_input_tokens",
            )
        }
    totals = {"calls": len(calls), "input_tokens": 0, "output_tokens": 0,
              "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0}
    for usage in calls.values():
        for key, value in usage.items():
            totals[key] += value
    return totals


def compute_metrics(
    outputs,
    answer_entries,
    file_problems=None,
    source_accruals=None,
    source_accrual_problems=None,
) -> dict[str, Any]:
    """Compute routing, control and finance metrics against the sealed key."""
    file_problems = list(file_problems or [])
    routed = [(outcome, record) for outcome, rows in outputs.items() for record in rows]
    routed_ids = [str(record.get("deduction_id")) for _, record in routed]
    id_counts = Counter(routed_ids)
    duplicate_ids = sorted(key for key, count in id_counts.items() if count > 1)
    unknown_ids = sorted(set(routed_ids) - set(answer_entries))
    omitted_ids = sorted(set(answer_entries) - set(routed_ids))

    confusion = {
        seed: {outcome: 0 for outcome in OUTCOME_FILES}
        for seed in ("auto_matched", "classifiable", "unresolvable")
    }
    for outcome, record in routed:
        truth = answer_entries.get(str(record.get("deduction_id")))
        if truth and truth.get("seed_category") in confusion:
            confusion[truth["seed_category"]][outcome] += 1

    auto_rows = outputs["auto_matched"]
    class_rows = outputs["classified_verified"]
    unres_rows = outputs["unresolvable"]
    seeded_auto = sum(e["seed_category"] == "auto_matched" for e in answer_entries.values())
    seeded_class = sum(e["seed_category"] == "classifiable" for e in answer_entries.values())
    seeded_unres = sum(e["seed_category"] == "unresolvable" for e in answer_entries.values())

    auto_true = sum(
        answer_entries.get(str(row.get("deduction_id")), {}).get("seed_category") == "auto_matched"
        for row in auto_rows
    )
    pair_correct = sum(
        answer_entries.get(str(row.get("deduction_id")), {}).get("matched_accrual_id")
        == row.get("matched_accrual_id") for row in auto_rows
    )
    class_correct = sum(
        answer_entries.get(str(row.get("deduction_id")), {}).get("seed_category") == "classifiable"
        and answer_entries.get(str(row.get("deduction_id")), {}).get("true_bucket") == row.get("llm_bucket")
        for row in class_rows
    )
    unres_true = sum(
        answer_entries.get(str(row.get("deduction_id")), {}).get("seed_category") == "unresolvable"
        for row in unres_rows
    )

    terminal_ids = []
    for outcome, row in routed:
        deduction_id = str(row.get("deduction_id"))
        truth = answer_entries.get(deduction_id)
        if not truth or outcome in {"needs_review", "data_quality_issue"}:
            continue
        expected = {"auto_matched": "auto_matched", "classifiable": "classified_verified",
                    "unresolvable": "unresolvable"}[truth["seed_category"]]
        wrong_bucket = outcome == "classified_verified" and row.get("llm_bucket") != truth.get("true_bucket")
        if outcome != expected or wrong_bucket:
            terminal_ids.append(deduction_id)
    terminal_ids = sorted(set(terminal_ids))

    source_accrual_problems = list(source_accrual_problems or [])
    source_by_id, source_id_counts = {}, Counter()
    for row in source_accruals or []:
        accrual_id = str(row.get("accrual_id") or "").strip()
        source_id_counts[accrual_id] += 1
        if accrual_id and accrual_id not in source_by_id:
            source_by_id[accrual_id] = row
    if source_accruals is not None:
        if source_id_counts.get(""):
            source_accrual_problems.append("source_accruals:blank accrual_id")
        duplicate_source_ids = sorted(
            identifier for identifier, count in source_id_counts.items()
            if identifier and count > 1
        )
        if duplicate_source_ids:
            source_accrual_problems.append(
                f"source_accruals:duplicate accrual_id count={len(duplicate_source_ids)}"
            )
    allocations_by_id = defaultdict(int)
    transitions_by_id = defaultdict(list)
    unknown_evidence_ids = set()
    invalid_evidence_allocation_ids = set()
    invalid_evidence_semantic_ids = set()
    candidate_source_bound_evidence_ids = set()
    deductions_by_evidence_id = defaultdict(set)
    matched_source_counts = Counter()
    invalid_transaction_match_ids = set()
    for row in auto_rows:
        accrual_id = str(row.get("matched_accrual_id") or "").strip()
        matched_source_counts[accrual_id] += 1
        source = source_by_id.get(accrual_id)
        if not accrual_id or source is None or source.get("evidence_scope") != "transaction_match":
            invalid_transaction_match_ids.add(accrual_id or "<blank>")
            continue
        try:
            pair_valid = (
                abs(to_cents(row.get("amount")) - to_cents(source.get("amount"))) <= 1
                and days_between(
                    row.get("_normalised_date") or row.get("transaction_date"),
                    source.get("accrual_date"),
                ) <= 1
                and vendor_score(
                    row.get("_normalised_vendor") or row.get("vendor_name"),
                    source.get("vendor_name"),
                ) >= 60
                and row.get("matched_bucket") == source.get("bucket")
            )
        except (TypeError, ValueError, OverflowError):
            pair_valid = False
        if not pair_valid:
            invalid_transaction_match_ids.add(accrual_id)
    reused_transaction_match_ids = sorted(
        identifier for identifier, count in matched_source_counts.items()
        if identifier and count > 1
    )
    transaction_match_integrity = (
        source_accruals is not None
        and not source_accrual_problems
        and not invalid_transaction_match_ids
        and not reused_transaction_match_ids
    )
    for row in class_rows:
        deduction_id = str(row.get("deduction_id") or "").strip()
        accrual_id = str(row.get("_evidence_accrual_id") or "").strip()
        if not accrual_id or accrual_id not in source_by_id:
            unknown_evidence_ids.add(accrual_id or "<blank>")
            continue
        source = source_by_id[accrual_id]
        deductions_by_evidence_id[accrual_id].add(deduction_id)
        if source.get("evidence_scope") != "programme_pool":
            invalid_evidence_allocation_ids.add(accrual_id)
            continue
        truth = answer_entries.get(deduction_id, {})
        expected_support = str(truth.get("supporting_accrual_id") or "").strip()
        try:
            semantic_evidence_valid = (
                truth.get("seed_category") == "classifiable"
                and expected_support == accrual_id
                and source.get("bucket") == row.get("llm_bucket")
                and vendor_score(
                    row.get("vendor_name"),
                    source.get("vendor_name"),
                ) >= 60
                and days_between(
                    row.get("transaction_date"),
                    source.get("accrual_date"),
                ) <= 45
            )
        except (TypeError, ValueError, OverflowError):
            semantic_evidence_valid = False
        if not semantic_evidence_valid:
            invalid_evidence_semantic_ids.add(deduction_id or "<blank>")
        deduction_cents = round(record_value(row) * 100)
        allocation_valid = True
        try:
            allocated = float(row.get("_allocated_amount"))
            before = float(row.get("_evidence_balance_before"))
            after = float(row.get("_evidence_balance_after"))
            if not all(math.isfinite(value) for value in (allocated, before, after)):
                raise ValueError("non-finite balance")
            allocated_cents = round(allocated * 100)
            before_cents = round(before * 100)
            after_cents = round(after * 100)
        except (TypeError, ValueError):
            invalid_evidence_allocation_ids.add(accrual_id)
            allocation_valid = False
            continue
        if (
            allocated_cents <= 0
            or allocated_cents != deduction_cents
            or after_cents < 0
            or before_cents - after_cents != allocated_cents
        ):
            invalid_evidence_allocation_ids.add(accrual_id)
            allocation_valid = False
        if semantic_evidence_valid and allocation_valid:
            candidate_source_bound_evidence_ids.add(deduction_id)
        allocations_by_id[accrual_id] += allocated_cents
        transitions_by_id[accrual_id].append((before_cents, after_cents, allocated_cents))
    overallocated_accrual_ids = []
    for accrual_id, allocated_cents in allocations_by_id.items():
        opening_cents = round(record_value(source_by_id[accrual_id]) * 100)
        if opening_cents <= 0 or allocated_cents > opening_cents:
            overallocated_accrual_ids.append(accrual_id)
            continue
        remaining = list(transitions_by_id[accrual_id])
        current = opening_cents
        while remaining:
            candidates = [entry for entry in remaining if entry[0] == current]
            if len(candidates) != 1:
                invalid_evidence_allocation_ids.add(accrual_id)
                break
            transition = candidates[0]
            remaining.remove(transition)
            current = transition[1]
    invalid_source_ids = set(invalid_evidence_allocation_ids) | set(overallocated_accrual_ids)
    valid_source_bound_evidence_ids = candidate_source_bound_evidence_ids - {
        deduction_id
        for accrual_id in invalid_source_ids
        for deduction_id in deductions_by_evidence_id.get(accrual_id, set())
    }
    evidence_complete = len(valid_source_bound_evidence_ids)
    source_allocation_integrity = (
        source_accruals is not None
        and not source_accrual_problems
        and not unknown_evidence_ids
        and not invalid_evidence_allocation_ids
        and not invalid_evidence_semantic_ids
        and not overallocated_accrual_ids
        and len(valid_source_bound_evidence_ids) == len(class_rows)
    )

    outcome_counts = {outcome: len(rows) for outcome, rows in outputs.items()}
    outcome_values = {outcome: round(sum(record_value(r) for r in rows), 2)
                      for outcome, rows in outputs.items()}
    total_value = round(sum(outcome_values.values()), 2)
    allocated_classified_value = round(sum(
        record_value(row) for row in class_rows
        if str(row.get("deduction_id") or "") in valid_source_bound_evidence_ids
    ), 2)
    resolved_value = round(outcome_values["auto_matched"] + allocated_classified_value, 2)
    terminal_set = set(terminal_ids)
    terminal_value = round(sum(record_value(row) for _, row in routed
                               if str(row.get("deduction_id")) in terminal_set), 2)
    accepted_method_counts = Counter(
        str(row.get("_classification_method") or "not_recorded") for row in class_rows
    )
    unresolvable_method_counts = Counter(
        str(row.get("_classification_method") or "not_recorded") for row in unres_rows
    )

    metrics = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "expected_lines": len(answer_entries),
        "routed_records": len(routed),
        "unique_routed_lines": len(set(routed_ids)),
        "file_problems": file_problems,
        "duplicate_ids": duplicate_ids,
        "unknown_ids": unknown_ids,
        "omitted_ids": omitted_ids,
        "outcome_counts": outcome_counts,
        "outcome_values_gbp": outcome_values,
        "total_value_gbp": total_value,
        "resolved_value_gbp": resolved_value,
        "allocated_classified_value_gbp": allocated_classified_value,
        "v1_plausibility_only_value_gbp": round(outcome_values["classified_verified"] - allocated_classified_value, 2),
        "human_work_value_gbp": round(total_value - resolved_value, 2),
        "resolved_value_rate": rate(round(resolved_value * 100), round(total_value * 100)),
        "auto_match_precision": rate(auto_true, len(auto_rows)),
        "auto_match_recall": rate(auto_true, seeded_auto),
        "auto_match_pair_correctness": rate(pair_correct, len(auto_rows)),
        "classification_precision": rate(class_correct, len(class_rows)),
        "accepted_classification_method_counts": dict(sorted(accepted_method_counts.items())),
        "unresolvable_method_counts": dict(sorted(unresolvable_method_counts.items())),
        "classifiable_automation_coverage": rate(class_correct, seeded_class),
        "evidence_allocation_coverage": rate(evidence_complete, len(class_rows)),
        "transaction_match_integrity": transaction_match_integrity,
        "invalid_transaction_match_ids": sorted(invalid_transaction_match_ids),
        "reused_transaction_match_ids": reused_transaction_match_ids,
        "source_allocation_integrity": source_allocation_integrity,
        "source_accrual_problems": sorted(set(source_accrual_problems)),
        "unknown_evidence_ids": sorted(unknown_evidence_ids),
        "invalid_evidence_allocation_ids": sorted(invalid_evidence_allocation_ids),
        "invalid_evidence_semantic_ids": sorted(invalid_evidence_semantic_ids),
        "valid_source_bound_evidence_ids": sorted(valid_source_bound_evidence_ids),
        "overallocated_accrual_ids": sorted(overallocated_accrual_ids),
        "unresolvable_precision": rate(unres_true, len(unres_rows)),
        "unresolvable_recall": rate(unres_true, seeded_unres),
        "terminal_misroute_count": len(terminal_ids),
        "terminal_misroute_ids": terminal_ids,
        "terminal_misroute_value_gbp": terminal_value,
        "value_weighted_terminal_misroute_rate": rate(round(terminal_value * 100), round(total_value * 100)),
        "confusion_matrix": confusion,
        "token_totals": dedupe_token_totals(
            class_rows + outputs["needs_review"] + unres_rows
        ),
    }

    incomplete = bool(
        file_problems or source_accrual_problems or duplicate_ids or unknown_ids or omitted_ids
    )
    if not routed:
        status, reason = "Not Assessable", "No routed records were available."
    elif incomplete:
        status = "Incomplete Run"
        reason = "Expected outputs or routing coverage are incomplete; no positive conclusion is allowed."
    else:
        gates = [
            metrics["terminal_misroute_count"] == 0,
            (metrics["auto_match_pair_correctness"] or 0) >= 0.99,
            metrics["transaction_match_integrity"],
            (metrics["classification_precision"] or 0) >= 0.95,
            (metrics["evidence_allocation_coverage"] or 0) >= 1.0,
            metrics["source_allocation_integrity"],
            (metrics["unresolvable_precision"] or 0) >= 0.95,
            (metrics["unresolvable_recall"] or 0) >= 0.95,
        ]
        status = "Ready for Demo" if all(gates) else "Needs Repair"
        reason = (
            "The complete synthetic run passed every published safety and quality gate."
            if status == "Ready for Demo" else
            "The run is complete, but one or more routing, precision or evidence gates failed."
        )
    metrics["overall_status"] = status
    metrics["overall_status_reason"] = reason
    return metrics


def priority_for(record, outcome) -> str:
    value = record_value(record)
    if (outcome in {"unresolvable", "data_quality_issue", "evidence_gap"} and value >= 250) or value >= 1000:
        return "P1"
    return "P2" if value >= 250 else "P3"


def next_action_for(outcome) -> str:
    return {
        "auto_matched": "No matching review required; any accounting action remains human-controlled.",
        "classified_verified": "No classification review required; any accounting action remains human-controlled.",
        "needs_review": "Confirm or correct the proposed bucket and supporting accrual.",
        "unresolvable": "Request remittance or portal evidence; do not post or write off automatically.",
        "data_quality_issue": "Correct the source field shown in the error column, then rerun.",
        "evidence_gap": "Rerun through V2 balance allocation; do not treat the V1 bucket check as evidence.",
    }.get(outcome, "Review the evidence before taking any accounting action.")


def excel_safe(value: Any) -> Any:
    """Keep source and model-controlled text inert in spreadsheet clients."""
    if not isinstance(value, str):
        return value
    stripped = value.lstrip(" \t\r\n")
    if (stripped and stripped[0] in "=+-@") or value.startswith(("\t", "\r")):
        return "'" + value
    return value


def write_header(ws, headers, row=1) -> None:
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row, column, header)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)
    ws.freeze_panes = ws.cell(row + 1, 1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def autosize(ws, minimum=11, maximum=55) -> None:
    widths = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(widths[cell.column], len(str(cell.value)))
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, minimum), maximum)


BASE_COLUMNS = [
    ("deduction_id", "Deduction ID"), ("transaction_date", "Date"),
    ("vendor_name", "Counterparty"), ("amount", "Deduction Value"),
    ("reference_code", "Reference"), ("description", "Description"),
]
AUTO_COLUMNS = BASE_COLUMNS + [
    ("matched_accrual_id", "Matched Accrual"), ("matched_bucket", "Bucket"),
    ("vendor_score", "Vendor Score"), ("date_delta_days", "Date Difference"),
    ("match_reason", "Evidence"), ("_display_status", "Status"),
]
CLASSIFIED_COLUMNS = BASE_COLUMNS + [
    ("llm_bucket", "Proposed Bucket"), ("_classification_method", "Classification Method"),
    ("llm_confidence", "Proposal Confidence"),
    ("llm_reasoning", "Classification Rationale"), ("_evidence_accrual_id", "Allocated Accrual"),
    ("_allocated_amount", "Allocated Amount"),
    ("_evidence_balance_before", "Balance Before"),
    ("_evidence_balance_after", "Balance After"),
    ("_evidence_note", "Control Evidence"), ("_display_status", "Status"),
    ("_display_next_action", "Next Action"),
]
EXCEPTION_COLUMNS = [("_display_priority", "Priority")] + BASE_COLUMNS + [
    ("llm_bucket", "Proposed Bucket"), ("llm_confidence", "AI Confidence"),
    ("_classification_error", "System Error"), ("_evidence_note", "Why It Stopped"),
    ("_display_owner", "Owner"), ("_display_status", "Status"),
    ("_display_next_action", "Next Action"),
]


def write_records_tab(ws, records, columns, outcome, valid_evidence_ids=None) -> None:
    valid_evidence_ids = set(valid_evidence_ids or [])
    write_header(ws, [label for _, label in columns])
    if not records:
        empty_messages = {
            "auto_matched": "No deductions were deterministically matched in this run.",
            "classified_verified": "No classifications passed the evidence controls in this run.",
            "needs_review": "No deductions need human classification review in this run.",
            "unresolvable": "No deductions were unresolvable from the supplied data in this run.",
            "data_quality_issue": "No source-data quality issues were detected in this run.",
        }
        autosize(ws)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        cell = ws.cell(2, 1, empty_messages.get(outcome, "No records were routed to this outcome in this run."))
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(italic=True, color=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28
        return
    for row_index, record in enumerate(sorted(records, key=record_value, reverse=True), 2):
        source_bound_evidence_valid = (
            outcome != "classified_verified"
            or str(record.get("deduction_id") or "") in valid_evidence_ids
        )
        enriched = {
            **record,
            "_display_priority": priority_for(record, outcome),
            "_display_owner": record.get("owner") or "Finance analyst",
            "_display_status": record.get("status") or (
                "Evidence gap" if not source_bound_evidence_valid
                else "Matched — deterministic" if outcome == "auto_matched"
                else "Accepted — evidence allocated" if outcome == "classified_verified" else "Open"
            ),
            "_display_next_action": record.get("next_action") or (
                next_action_for("evidence_gap")
                if not source_bound_evidence_valid
                else next_action_for(outcome)
            ),
        }
        for column_index, (key, _label) in enumerate(columns, 1):
            cell = ws.cell(row_index, column_index, excel_safe(enriched.get(key)))
            if key in {"amount", "_allocated_amount", "_evidence_balance_before", "_evidence_balance_after"}:
                cell.number_format = '"£"#,##0.00;[Red]-"£"#,##0.00'
            if key in {"description", "match_reason", "llm_reasoning", "_evidence_note", "_display_next_action"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)


def build_summary(ws, metrics, outputs) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Deduction Resolution Workbench"
    ws["A1"].font = Font(size=20, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 32
    ws["A2"] = "Synthetic demonstration only — no real client data; Ready for Demo is not production approval."
    ws["A2"].font = Font(italic=True, color="44546A")

    ws["A4"], ws["B4"] = "RUN DECISION", metrics["overall_status"]
    ws["A4"].font, ws["A4"].fill = Font(bold=True, color=WHITE), PatternFill("solid", fgColor=BLUE)
    ws["B4"].font, ws["B4"].fill = Font(bold=True, size=13), STATUS_FILLS[metrics["overall_status"]]
    ws["C4"] = metrics["overall_status_reason"]
    ws["C4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[4].height = 32

    ws["A6"], ws["A6"].fill = "FINANCE OUTCOME — START HERE", SECTION_FILL
    ws["A6"].font = Font(bold=True, color=NAVY)
    cards = [
        ("Total deductions", metrics["total_value_gbp"], NAVY, '"£"#,##0.00'),
        ("Controlled resolved value", metrics["resolved_value_gbp"], GREEN, '"£"#,##0.00'),
        ("Value not yet controlled", metrics["human_work_value_gbp"], RED, '"£"#,##0.00'),
        ("Terminal routing errors", metrics["terminal_misroute_count"],
         GREEN if metrics["terminal_misroute_count"] == 0 else RED, '0'),
    ]
    for column, (label, value, colour, number_format) in enumerate(cards, 1):
        ws.cell(8, column, label).font = Font(bold=True, color=WHITE)
        ws.cell(8, column).fill = PatternFill("solid", fgColor=colour)
        ws.cell(8, column).alignment = Alignment(horizontal="center", wrap_text=True)
        cell = ws.cell(9, column, value)
        cell.font, cell.number_format = Font(size=18, bold=True, color=colour), number_format
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[9].height = 30

    ws["A12"], ws["A12"].fill = "WHAT THE FINANCE TEAM SHOULD DO", SECTION_FILL
    ws["A12"].font = Font(bold=True, color=NAVY)
    open_work = [(priority_for(r, o), record_value(r), o, r)
                 for o in ("needs_review", "unresolvable", "data_quality_issue")
                 for r in outputs[o]]
    open_work.extend(
        (priority_for(r, "evidence_gap"), record_value(r), "evidence_gap", r)
        for r in outputs["classified_verified"] if not has_reconciling_allocation(r)
    )
    open_work.sort(key=lambda item: (item[0], -item[1]))
    if open_work:
        _, value, outcome, record = open_work[0]
        ws["A14"] = f"First investigate {record.get('deduction_id')} (£{value:,.2f}): {next_action_for(outcome)}"
    else:
        ws["A14"] = (
            "No classification or data-quality investigation is required by this synthetic run; "
            "accounting actions remain human-controlled."
        )
    ws["A14"].font, ws["A14"].alignment = Font(bold=True), Alignment(wrap_text=True)

    write_header(ws, ["Outcome", "Lines", "Value", "Share of value"], row=17)
    for row_index, outcome in enumerate(OUTCOME_FILES, 18):
        value = metrics["outcome_values_gbp"][outcome]
        ws.cell(row_index, 1, OUTCOME_LABELS[outcome])
        ws.cell(row_index, 2, metrics["outcome_counts"][outcome])
        ws.cell(row_index, 3, value).number_format = '"£"#,##0.00'
        ws.cell(row_index, 4, rate(round(value * 100), round(metrics["total_value_gbp"] * 100)) or 0).number_format = "0.0%"
    chart = DoughnutChart()
    chart.title = "Where the deduction value went"
    chart.add_data(Reference(ws, min_col=3, min_row=17, max_row=22), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=18, max_row=22))
    chart.height, chart.width = 7, 11
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    ws.add_chart(chart, "F17")

    ws["A25"], ws["A25"].fill = "TRUST AND CONTROL CHECKS", SECTION_FILL
    ws["A25"].font = Font(bold=True, color=NAVY)
    write_header(ws, ["Control", "Result", "What it means"], row=27)
    controls = [
        ("Auto-match pair correctness", metrics["auto_match_pair_correctness"], "Were deterministic pairs correct?"),
        ("Transaction-match source integrity", 1.0 if metrics["transaction_match_integrity"] else 0.0, "Does every auto-match use one known transaction accrual exactly once?"),
        ("Classification precision", metrics["classification_precision"], "Were accepted classifications correct, regardless of whether they came from a configured alias or an AI proposal?"),
        ("Unresolvable precision", metrics["unresolvable_precision"], "Of lines called unresolvable, how many truly were?"),
        ("Unresolvable recall", metrics["unresolvable_recall"], "Of truly unresolvable lines, how many were caught?"),
        ("Classifiable automation coverage", metrics["classifiable_automation_coverage"], "How much classifiable work was safely automated?"),
        ("Allocated-evidence coverage", metrics["evidence_allocation_coverage"], "Did every accepted classification consume a sufficient balance?"),
        ("Source-ledger allocation integrity", 1.0 if metrics["source_allocation_integrity"] else 0.0, "Do aggregate allocations stay within one unique source balance?"),
    ]
    for row_index, (label, result, note) in enumerate(controls, 28):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, pct(result))
        ws.cell(row_index, 3, note)
    ws["A38"], ws["A38"].fill = "CLASSIFICATION ROUTE DISCLOSURE", SECTION_FILL
    ws["A38"].font = Font(bold=True, color=NAVY)
    write_header(ws, ["Route", "Lines", "Interpretation"], row=40)
    disclosure_rows = [
        (
            "Accepted via configured aliases",
            metrics["accepted_classification_method_counts"].get("configured_alias", 0),
            "Deterministic finance-team configuration; not AI performance.",
        ),
        (
            "Accepted via AI proposals",
            metrics["accepted_classification_method_counts"].get("ai_proposal", 0),
            "AI-proposed buckets that passed the disclosed evidence controls.",
        ),
        (
            "AI abstentions routed as unresolvable",
            metrics["unresolvable_method_counts"].get("ai_proposal", 0),
            "AI proposals of Unresolvable that passed the conflict check.",
        ),
    ]
    for row_index, (route, count, interpretation) in enumerate(disclosure_rows, 41):
        ws.cell(row_index, 1, route)
        ws.cell(row_index, 2, count)
        ws.cell(row_index, 3, interpretation)
    widths = {"A": 38, "B": 24, "C": 44, "D": 24, "E": 4, "F": 18, "G": 18, "H": 18}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_priority_worklist(ws, outputs, valid_evidence_ids=None) -> None:
    valid_evidence_ids = set(valid_evidence_ids or [])
    rows = [(outcome, record) for outcome in ("needs_review", "unresolvable", "data_quality_issue")
            for record in outputs[outcome]]
    rows.extend(
        ("evidence_gap", record) for record in outputs["classified_verified"]
        if str(record.get("deduction_id") or "") not in valid_evidence_ids
    )
    rows.sort(key=lambda item: (priority_for(item[1], item[0]), -record_value(item[1])))
    write_header(ws, ["Priority", "Outcome", "Deduction ID", "Date", "Counterparty", "Value",
                      "Proposed Bucket", "Why It Stopped", "Owner", "Status", "Next Action"])
    for row_index, (outcome, record) in enumerate(rows, 2):
        values = [
            priority_for(record, outcome), (
                "Classification — evidence gap" if outcome == "evidence_gap" else OUTCOME_LABELS[outcome]
            ), record.get("deduction_id"),
            record.get("transaction_date"), record.get("vendor_name"), record_value(record),
            record.get("llm_bucket"), (
                "The proposed bucket did not receive a reconciling balance allocation."
                if outcome == "evidence_gap" else
                record.get("_evidence_note") or record.get("error") or record.get("_classification_error")
            ),
            record.get("owner") or "Finance analyst", record.get("status") or "Open", next_action_for(outcome),
        ]
        for column_index, value in enumerate(values, 1):
            cell = ws.cell(row_index, column_index, excel_safe(value))
            if column_index == 6:
                cell.number_format = '"£"#,##0.00'
            if column_index in {8, 11}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor={"P1": LIGHT_RED, "P2": LIGHT_AMBER}.get(values[0], GREY))
    autosize(ws)


def build_evaluation_tab(ws, metrics) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Evaluation — full routing evidence"
    ws["A1"].font = Font(size=16, bold=True, color=NAVY)
    ws["A3"] = "Every percentage is measured only against this synthetic answer key; it is not production accuracy."
    ws["A3"].font = Font(italic=True, color=RED)
    write_header(ws, ["Seeded truth", *[OUTCOME_LABELS[o] for o in OUTCOME_FILES], "Total"], row=5)
    for row_index, seed in enumerate(("auto_matched", "classifiable", "unresolvable"), 6):
        ws.cell(row_index, 1, seed.replace("_", " ").title())
        for column_index, outcome in enumerate(OUTCOME_FILES, 2):
            ws.cell(row_index, column_index, metrics["confusion_matrix"][seed][outcome])
        ws.cell(row_index, 7, f"=SUM(B{row_index}:F{row_index})")
    write_header(ws, ["Metric", "Result", "Numerator / denominator meaning"], row=12)
    rows = [
        ("Auto-match precision", metrics["auto_match_precision"], "True matchable lines / all lines auto-matched"),
        ("Auto-match recall", metrics["auto_match_recall"], "Matchable lines auto-matched / all seeded matchable lines"),
        ("Pair correctness", metrics["auto_match_pair_correctness"], "Correct accrual pairs / all auto-matched pairs"),
        ("Transaction-match source integrity", 1.0 if metrics["transaction_match_integrity"] else 0.0, "Every auto-match references one known transaction accrual exactly once"),
        ("Classification precision", metrics["classification_precision"], "Correct accepted classifications / all accepted classifications"),
        ("Accepted via configured aliases", metrics["accepted_classification_method_counts"].get("configured_alias", 0), "Deterministic configured classifications; count, not a percentage"),
        ("Accepted via AI proposals", metrics["accepted_classification_method_counts"].get("ai_proposal", 0), "AI-proposed classifications accepted after evidence controls; count, not a percentage"),
        ("AI abstentions routed as unresolvable", metrics["unresolvable_method_counts"].get("ai_proposal", 0), "AI proposals of Unresolvable accepted after the conflict check; count, not a percentage"),
        ("Classifiable automation coverage", metrics["classifiable_automation_coverage"], "Correct accepted classifications / all seeded classifiable lines"),
        ("Allocated-evidence coverage", metrics["evidence_allocation_coverage"], "Accepted lines with a reconciling non-negative balance allocation / all accepted classifications"),
        ("Source-ledger allocation integrity", 1.0 if metrics["source_allocation_integrity"] else 0.0, "Unique source IDs, known evidence IDs and aggregate allocations within opening balances"),
        ("Unresolvable precision", metrics["unresolvable_precision"], "Truly unresolvable / all lines labelled unresolvable"),
        ("Unresolvable recall", metrics["unresolvable_recall"], "Truly unresolvable caught / all seeded unresolvable lines"),
        ("Value-weighted terminal error rate", metrics["value_weighted_terminal_misroute_rate"], "Value in unsafe terminal routes / total routed value"),
    ]
    for row_index, (label, result, meaning) in enumerate(rows, 13):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, result if isinstance(result, int) else pct(result))
        ws.cell(row_index, 3, meaning)
    ws["A29"], ws["A29"].font = "Status logic", Font(bold=True)
    for row, (status, meaning) in enumerate([
        ("Incomplete Run", "Any expected branch/source file is missing or invalid, or an input/source ID is duplicate, unknown or omitted."),
        ("Needs Repair", "Complete run, but a terminal misroute or safety/precision failure exists."),
        ("Ready for Demo", "Complete run, zero terminal misroutes and every published gate passes."),
    ], 30):
        ws.cell(row, 1, status)
        ws.cell(row, 2, meaning)
    autosize(ws)


def build_workbook(outputs, metrics, destination: Path) -> None:
    wb = Workbook()
    start = wb.active
    start.title = "Start_Here"
    build_summary(start, metrics, outputs)
    valid_evidence_ids = metrics.get("valid_source_bound_evidence_ids", [])
    build_priority_worklist(
        wb.create_sheet("Priority_Worklist"), outputs, valid_evidence_ids
    )
    write_records_tab(wb.create_sheet("Auto_Matched"), outputs["auto_matched"], AUTO_COLUMNS, "auto_matched")
    write_records_tab(
        wb.create_sheet("Classified_Evidence"), outputs["classified_verified"],
        CLASSIFIED_COLUMNS, "classified_verified", valid_evidence_ids,
    )
    write_records_tab(wb.create_sheet("Needs_Review"), outputs["needs_review"], EXCEPTION_COLUMNS, "needs_review")
    write_records_tab(wb.create_sheet("Unresolvable"), outputs["unresolvable"], EXCEPTION_COLUMNS, "unresolvable")
    write_records_tab(wb.create_sheet("Data_Quality"), outputs["data_quality_issue"], EXCEPTION_COLUMNS, "data_quality_issue")
    build_evaluation_tab(wb.create_sheet("Evaluation"), metrics)
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--outputs-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--answer-key", type=Path, default=DEFAULT_ANSWER_KEY)
    parser.add_argument(
        "--accrual-csv",
        type=Path,
        default=None,
        help="Authoritative accrual ledger required for a source-bound Ready for Demo result.",
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--metrics-json", type=Path, default=DEFAULT_METRICS_JSON)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    outputs, file_problems = load_run_outputs(args.outputs_dir)
    _seed_counts, entries = load_answer_key(args.answer_key)
    if args.accrual_csv is None:
        source_accruals, source_accrual_problems = None, []
    else:
        source_accruals, source_accrual_problems = load_source_accruals(args.accrual_csv)
    metrics = compute_metrics(
        outputs,
        entries,
        file_problems,
        source_accruals,
        source_accrual_problems,
    )
    build_workbook(outputs, metrics, args.xlsx)
    args.metrics_json.parent.mkdir(parents=True, exist_ok=True)
    args.metrics_json.write_text(json.dumps(metrics, indent=2) + "\n", encoding="utf-8")
    print(f"Workbook: {args.xlsx}")
    print(f"Evaluation: {args.metrics_json}")
    print(f"Overall status: {metrics['overall_status']}")
    print(f"Unresolvable precision: {pct(metrics['unresolvable_precision'])}")
    print(f"Unresolvable recall: {pct(metrics['unresolvable_recall'])}")
    print(f"Terminal misroutes: {metrics['terminal_misroute_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
